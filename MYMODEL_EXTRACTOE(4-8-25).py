import pdfplumber
import pandas as pd
import re
import joblib
import openpyxl

# ========== Load your trained classification model and vectorizer ==========
Model = joblib.load("MODELS/ROG_CLASSIFIER.pkl")
vectorizer = joblib.load("MODELS/ROG_tfidf_vectorizer.pkl")

def preprocess_text(text):
    text = text.lower()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    return text

# ========== PDF and output config ==========
pdf_path = "TOI.pdf"  # Your input PDF path
output_excel = "TOOOOOOOOOOOOOOOOOOOOOI.xlsx"  # Your output Excel file

# Regex patterns used for parsing financial lines
NOTE_RE = re.compile(r'^\d+([a-zA-Z])?(\(\w+\))?(\.\d+)?([a-zA-Z]?)?$|^\d+\.\d+(\([a-zA-Z0-9]+\))?$|^\d+\([a-zA-Z0-9]+\)$')
VALUE_RE = re.compile(r'^\(?-?[\d,]+(?:\.\d+)?\)?$|^-$')

def find_years_in_text(lines):
    year_candidates = []
    for line in lines[:15]:
        found = re.findall(r"(?:as at\s*)?([A-Za-z]+\s*\d{1,2},\s*\d{4}|\d{1,2}\s+[A-Za-z]+\s+\d{4}|[A-Za-z]+\s+\d{4})", line, re.I)
        for y in found:
            y_clean = re.sub("^as at\s*", "", y, flags=re.I).strip()
            if y_clean and y_clean not in year_candidates:
                year_candidates.append(y_clean)
        if len(year_candidates) >= 2:
            break
    if len(year_candidates) < 2:
        for line in lines[:12]:
            found = re.findall(r"(?:as at\s*)?([A-Za-z]+\s*\d{1,2},\s*\d{4}|\d{1,2}\s+[A-Za-z]+\s+\d{4}|[A-Za-z]+\s+\d{4})", line, re.I)
            for y in found:
                y_clean = re.sub("^as at\s*", "", y, flags=re.I).strip()
                if y_clean and y_clean not in year_candidates:
                    year_candidates.append(y_clean)
            if len(year_candidates) >= 2:
                break
    if len(year_candidates) >= 2:
        return year_candidates[:2]
    return ["Year1", "Year2"]

def parse_financial_lines(lines, year_cols, cashflow_section=False):
    data = []
    for line in lines:
        line = re.sub(r'\s+', ' ', line.strip())
        if not line:
            continue
        tokens = line.split(' ')
        n = len(tokens)
        particulars = ''
        note = ''
        year1 = ''
        year2 = ''
        if n >= 2 and VALUE_RE.fullmatch(tokens[-1]) and VALUE_RE.fullmatch(tokens[-2]):
            if n >= 3 and NOTE_RE.fullmatch(tokens[-3]):
                particulars = ' '.join(tokens[:-3])
                note = tokens[-3]
            else:
                particulars = ' '.join(tokens[:-2])
                note = ''
            year1 = tokens[-2]
            year2 = tokens[-1]
        elif n >= 1 and VALUE_RE.fullmatch(tokens[-1]):
            particulars = ' '.join(tokens[:-1])
            note = ''
            year1 = tokens[-1]
            year2 = ''
        elif n >= 1 and NOTE_RE.fullmatch(tokens[-1]):
            particulars = ' '.join(tokens[:-1])
            note = tokens[-1]
            year1 = ''
            year2 = ''
        else:
            particulars = line
            note = ''
            year1 = ''
            year2 = ''
        if cashflow_section:
            note = ''
        data.append({
            "Particulars": particulars.strip(),
            "Notes": note,
            year_cols[0]: year1,
            year_cols[1]: year2
        })
    return pd.DataFrame(data)

# ========== Label normalization map to align model output with dict keys ==========
label_map = {
    "Balance Sheets": "Balance Sheet",
    "Balance Sheet": "Balance Sheet",
    "Cash Flow": "Cash Flow Statement",
    "Cash Flow Statement": "Cash Flow Statement",
    "Income Statement": "Income Statement",
    "Notes": "Notes",
    "Others": "Others"
}

# ========== Main processing ==========

# Prepare dict to collect parsed data for each class (excluding "Others")
extracted_by_class = {
    "Balance Sheet": [],
    "Cash Flow Statement": [],
    "Income Statement": [],
    "Notes": []
}

with pdfplumber.open(pdf_path) as pdf:
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        if not text:
            continue

        # Preprocess and vectorize page text
        processed_text = preprocess_text(text)
        text_vec = vectorizer.transform([processed_text])

        # Predict raw document category then normalize it
        pred_class_raw = Model.predict(text_vec)[0]
        pred_class = label_map.get(pred_class_raw, "Others")  # Fallback to Others if unmapped

        # Skip Others category
        if pred_class == "Others":
            print(f"Page {i + 1}: Classified as '{pred_class}', skipped.")
            continue

        # Split text into lines
        lines = text.splitlines()

        # Extract years from header lines
        year_cols = find_years_in_text(lines)

        # Find index of header line containing 'PARTICULARS' and both year labels
        header_idx = None
        for idx, line in enumerate(lines):
            if "PARTICULARS" in line.upper() and all(y in line for y in year_cols):
                header_idx = idx
                break
        if header_idx is None:
            header_idx = 5  # Fallback index

        data_lines = lines[header_idx + 1:]

        # Check if cashflow section for parsing notes accordingly
        cashflow_section = (pred_class == "Cash Flow Statement")

        # Parse financial table lines into DataFrame
        parsed_df = parse_financial_lines(data_lines, year_cols, cashflow_section)

        # Collect parsed data by normalized predicted class with page info
        extracted_by_class[pred_class].append((f"Page {i + 1}", parsed_df))

        print(f"Page {i + 1}: Classified as '{pred_class}'")

# Export all parsed data to Excel file with one sheet per document category
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    for category, items in extracted_by_class.items():
        if not items:
            continue
        sheet_name = category[:31]  # Excel sheet name max length
        start_row = 0

        # Create or get worksheet to control writing
        if sheet_name in writer.book.sheetnames:
            worksheet = writer.book[sheet_name]
        else:
            worksheet = writer.book.create_sheet(sheet_name)

        for heading, df in items:
            # Write heading in bold font
            worksheet.cell(row=start_row + 1, column=1, value=heading).font = worksheet.cell(row=start_row + 1, column=1).font.copy(bold=True)
            # Write dataframe below heading
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=False, header=True)
            # Move start row to after current table plus spacing rows
            start_row += len(df) + 4

print(f"\nExtraction and classification complete! Data saved to '{output_excel}'")
